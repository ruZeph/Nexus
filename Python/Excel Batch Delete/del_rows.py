import openpyxl
from openpyxl.utils import get_column_letter
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import copy

def safe_delete_row(ws, row_idx):
    """
    Safely deletes a row by first checking if it is part of a merged cell.
    If it is the bottom of a merged cell (like Sl. No), it shrinks the merge
    instead of breaking it.
    """
    # Create a list of merges to modify (so we don't modify the list while iterating)
    merges_to_process = list(ws.merged_cells.ranges)
    
    for merge in merges_to_process:
        min_col, min_row, max_col, max_row = merge.bounds
        
        # Case 1: The merge is ENTIRELY inside the row we want to delete
        # (e.g., "TOTAL" cell might be merged B8:C8) -> Just unmerge it.
        if min_row == row_idx and max_row == row_idx:
            ws.unmerge_cells(str(merge))
            
        # Case 2: The merge ENDS on the row we want to delete
        # (e.g., Sl No spans Row 6 to Row 8. We are deleting Row 8.)
        elif min_row < row_idx and max_row == row_idx:
            # 1. Unmerge the original block
            ws.unmerge_cells(str(merge))
            
            # 2. Re-merge it, but make it 1 row shorter (stop at row_idx - 1)
            new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row - 1}"
            ws.merge_cells(new_range)

    # Now that merges are safe, delete the row normally
    ws.delete_rows(row_idx)

def process_excel_file():
    # 1. GUI Setup
    root = tk.Tk()
    root.withdraw()

    # 2. Select File
    input_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx;*.xlsm")]
    )

    if not input_path:
        return

    try:
        # 3. Output Path
        file_root, file_ext = os.path.splitext(input_path)
        output_path = f"{file_root}_fixed{file_ext}"

        print(f"Loading: {input_path}...")
        wb = openpyxl.load_workbook(input_path)
        
        # 4. Process Sheets
        target_sheets = wb.sheetnames[:8]

        for sheet_name in target_sheets:
            ws = wb[sheet_name]
            print(f"Scanning sheet: {sheet_name}")
            
            # Find rows containing "TOTAL" in Col A, B, or C
            rows_to_remove = []
            for row in ws.iter_rows():
                # Check first 3 columns safely
                val_a = str(row[0].value).strip().upper() if row[0].value else ""
                val_b = str(row[1].value).strip().upper() if len(row) > 1 and row[1].value else ""
                val_c = str(row[2].value).strip().upper() if len(row) > 2 and row[2].value else ""
                
                if "TOTAL" in val_a or "TOTAL" in val_b or "TOTAL" in val_c:
                    rows_to_remove.append(row[0].row)

            # SORT REVERSE is critical
            rows_to_remove.sort(reverse=True)
            
            for row_idx in rows_to_remove:
                print(f"  - Safely Deleting Row {row_idx}")
                safe_delete_row(ws, row_idx)

        # 5. Save
        wb.save(output_path)
        messagebox.showinfo("Success", f"Processing Complete!\n\nSaved as:\n{output_path}")

    except PermissionError:
        messagebox.showerror("Error", "Please close the Excel file before running this program.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred:\n{e}")

if __name__ == "__main__":
    process_excel_file()